import re, os
from ports_list import *
from openpyxl import Workbook

from functions import extract_ports_for_fw, create_new_worksheet, append_data_to_worksheet, get_lb_name, \
    validate_ip_address

wb = Workbook()


def palo_alto_extract():
    extract_ports_for_fw()
    pattern = r'\s+'

    with open('Palo Arp.txt') as arp_file:
        arp = arp_file.readlines()

        for line in arp:
            row = re.split(pattern, line)
            if row[0][:4] == 'ae1.':
                if row[2] == '(incomplete)':
                    continue
                data = [row[0], row[1], row[2]]

                if row[0] in internet_transfer:
                    if 'internet_transfer' not in wb.sheetnames:
                        create_new_worksheet('internet_transfer', wb)
                    append_data_to_worksheet(data, "internet_transfer", wb)

                elif row[0] in infra:
                    if 'infra' not in wb.sheetnames:
                        create_new_worksheet('infra', wb)
                    append_data_to_worksheet(data, "infra", wb)

                elif row[0] in msvx_nsx_t_dmz:
                    if 'msvx_nsx_t_dmz' not in wb.sheetnames:
                        create_new_worksheet('msvx_nsx_t_dmz', wb)
                    append_data_to_worksheet(data, 'msvx_nsx_t_dmz', wb)

                elif row[0] in msvx_nsx_t_default:
                    if 'msvx_nsx_t_default' not in wb.sheetnames:
                        create_new_worksheet('msvx_nsx_t_default', wb)
                    append_data_to_worksheet(data, "msvx_nsx_t_default", wb)

                elif row[0] in sap01:
                    if 'sap01' not in wb.sheetnames:
                        create_new_worksheet('sap01', wb)
                    append_data_to_worksheet(data, "sap01", wb)

                elif row[0] in deedvbasf005:
                    if 'deedvbasf005' not in wb.sheetnames:
                        create_new_worksheet('deedvbasf005', wb)
                    append_data_to_worksheet(data, "deedvbasf005", wb)

                elif row[0] in vmpchbi01:
                    if 'vmpchbi01' not in wb.sheetnames:
                        create_new_worksheet('vmpchbi01', wb)
                    append_data_to_worksheet(data, "vmpchbi01", wb)

                elif row[0] in vmpcsdcn01:
                    if 'vmpcsdcn01' not in wb.sheetnames:
                        create_new_worksheet('vmpcsdcn01', wb)
                    append_data_to_worksheet(data, "vmpcsdcn01", wb)

    wb.save('Firewalls ARP Table.xlsx')
    print(f"{'Firewalls'}: complete")


def check_point_extract():
    pattern = r'\s+'

    with open('CHECKPOINT_ARP_INVENTORY.txt') as arp_file:
        arp = arp_file.readlines()

        for line in arp:
            row = re.split(pattern, line)

            data = [row[0]]

            if row[0] == '?' and row[5] == 'on':
                if row[3] == '<incomplete>':
                    continue
                data = [row[6], row[1][1:-1], row[3]]
            elif row[0] == '?':
                if row[3] == '<incomplete>':
                    continue
                data = [row[5], row[1][1:-1], row[3]]

            if 'Check Point' not in wb.sheetnames:
                create_new_worksheet('Check Point', wb)
            append_data_to_worksheet(data, "Check Point", wb)

    wb.save('Firewalls ARP Table.xlsx')
    print("Check Point: complete")


def lb_extract():
    pattern = r'\s+'
    lb_directory = "LB_Files"
    for file in os.listdir(lb_directory):
        if file.endswith('.txt'):
            with open(os.path.join(lb_directory, file)) as f:
                arp_file = f.readlines()
                lb_name = get_lb_name(arp_file[0])

                for line in arp_file:
                    row = re.split(pattern, line)
                    if len(row) > 2 and row[2] == 'incomplete':
                        continue
                    ip = row[0].split('%')
                    if len(ip) == 2:
                        ip, partition = ip[0], ip[1]
                    else:
                        ip = str(row[0])

                    if not validate_ip_address(ip):
                        continue
                    vlan = re.split('_|#', row[3])[-1]
                    mac = row[2]
                    data = [vlan, ip, mac]

                    if lb_name not in wb.sheetnames:
                        create_new_worksheet(lb_name, wb)
                    append_data_to_worksheet(data, lb_name, wb)

            wb.save('Firewalls_And_LB ARP Table.xlsx')
            print(f"{lb_name}: complete")


palo_alto_extract()
check_point_extract()
lb_extract()
